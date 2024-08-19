"""
.txt
.json
.py
"""

import csv
import json

employees = [
    {
        "emp_name": 'Botir',
        'dept': 'Accounting',
        'birth_month': 'Noverber'
    },
    {
        "emp_name": 'Kamol',
        'dept': 'IT',
        'birth_month': 'March'
    }
]

# with open("employees.csv", mode = 'a') as f:
#     field_names = list(employees[0].keys())
#     writer = csv.DictWriter(f , field_names)
#     # writer.writeheader()
#     writer.writerows(employees)


# with open("employees.csv", mode = 'r') as f:
#     reader = csv.DictReader(f)
#     print(list(reader))

# with open("tasks/centers.json") as f:
#     centers = json.load(f)
#
# with open("center.csv" , 'w') as f:
#     fields = centers[0].keys()
#     writer = csv.DictWriter(f , fields)
#     writer.writeheader()
#     writer.writerows(centers)


# ===============================================================

import openpyxl # alt + enter + enter

# ================= read =============================
# wb_obj =openpyxl.load_workbook("tast2.xlsx")
# sheet_obj = wb_obj.active
# row = sheet_obj.max_row
# col = sheet_obj.max_column
#
# for i in range(1, row+1):
#     cell = sheet_obj.cell(i  , 1)
#     cemester = sheet_obj.cell(i  , 4)
#     print(cell.value , cemester.value)
# =========================== write =====================


# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet = wb.active
# c1 = sheet['A2']
# c1.value = 'hello'
#
# c2 = sheet['A1']
# c2.value = 'world'
#
# wb.save('sample.xlsx')









