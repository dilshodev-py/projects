import csv

# with open('employees.csv', mode='w') as csv_file:
#     header = ['emp_name' , "dept" , "birth_month"]
#     writer = csv.DictWriter(csv_file , fieldnames=header)
#
#     writer.writeheader()
#     # writer.writerow({'dept' : 'Accounting',"emp_name" : 'John Smith' , 'birth_month' : 'November'})
#     # writer.writerow({"emp_name" : 'Botir' , 'dept' : 'Accounting' , 'birth_month' : 'February'})
#     writer.writerows([{"emp_name" : 'Botir' , 'dept' : 'Accounting' , 'birth_month' : 'February'},{'dept' : 'Accounting',"emp_name" : 'John Smith' , 'birth_month' : 'November'}])

# with open('employees.csv', mode='r') as csv_file:
#     read = csv.DictReader(csv_file , delimiter=',')
#     print(list(read))


# =======================================================

# import json
#
# data = [
#     {
#         'id': 1,
#         "name": "Laptop"
#     },
#     {
#         'id': 2,
#         "name": "Laptop2"
#     }
# ]
# with open('data.json', 'w') as file:
#     json.dump(data , file , indent=3)
#
# with open('data.json', 'r') as file:
#     products = json.load(file)
# print(products)

import openpyxl

wb_obj = openpyxl.load_workbook('table.xlsx')
sheet_obj = wb_obj.active

max_row = sheet_obj.max_row
for i in range(1, max_row + 1):
    name = sheet_obj.cell(row=i, column=1).value
    semester = sheet_obj.cell(row=i, column=4).value
    print(f"{name} {semester}")


"""
Home work:
    cars.csv   ->    cars.json
    users.csv  ->    users.xlsx
    
    
    
"""
