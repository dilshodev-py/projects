import asyncio
import random

import openpyxl
import pandas as pd
from telethon import TelegramClient
from telethon import events
# python 2 4
# bilim 3


API_ID = 20219166
API_HASH = '0b023d8fb22abf1b5ee1db73be13373d'
phone_number = '+998981214545'

client = TelegramClient(f'session_{phone_number}', API_ID, API_HASH)


def from_specific_user(event):
    return event.text == 'send'
# C:/Users/User/Desktop/error_send.xlsx
path = "C:/Users/User/Desktop/"
@client.on(events.NewMessage(func=from_specific_user))
async def main(event):
    data = pd.read_excel(path + 'egg.xlsx')
    face_phone_number = []
    for i in data.values:
        try:
            message = i[1]
            if str(i[0])[1:].isdigit():
                user = await client.get_entity("+"+str(i[0]))
            else:
                user = await client.get_entity(str(i[0]))
            await client.send_message(user.id , message)
            await asyncio.sleep(random.randrange(4, 7))
        except:
            face_phone_number.append(i)
    with pd.ExcelWriter(path+"error_send.xlsx" , mode="w") as writer:
        pass
    wb_obj = openpyxl.load_workbook(path+"error_send.xlsx")
    sheet_obj = wb_obj.active
    sheet_obj.cell(1, 1, "number")
    sheet_obj.cell(1, 2, "message")
    sheet_obj.cell(1, 3, "name")
    for i , v in enumerate(face_phone_number,1):
        sheet_obj.cell(i, 1, v[0])
        sheet_obj.cell(i, 2, v[1])
        sheet_obj.cell(i, 3, v[2])
    wb_obj.save(path+"error_send.xlsx")


if __name__ == '__main__':
    client.start()
    client.run_until_disconnected()

