import csv
import json

# with open('centers.json') as f:
#     data = json.load(f)
#
# with open('centers.csv', 'w') as f:
#     fields_name = data[0].keys()
#     writer = csv.DictWriter(f, fields_name)
#     writer.writeheader(),writer.writerows(data)


# ======================

#
# with open('cars.csv') as f:
#     data = list(csv.DictReader(f))
#
#
# with open('cars.json', 'w') as f:
#     json.dump(data , f , indent=4)
#

import openpyxl


path = 'students.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet = wb_obj.active
max_row = sheet.max_row
for i in range(1, max_row + 1):
    name = sheet.cell(row=i, column=1).value
    semester = sheet.cell(row=i, column=4).value
    print(name , semester)





