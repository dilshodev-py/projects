import datetime
from openpyxl import Workbook , load_workbook

# wb = Workbook()
# ws = wb.active

# ws['A1'] = 42
# ws['B1'] = datetime.datetime.now()
# wb.save("sample.xlsx")

path = "/home/dilshod/PycharmProject/p18_group/module_3/lesson_8/sample.xlsx"

wb_obj = load_workbook(path)

sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=2)
for i in range(1 , sheet_obj.max_row +1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)



