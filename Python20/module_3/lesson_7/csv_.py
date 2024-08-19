import csv

# March


# ========================= read ===========================
# with open('users.csv') as csv_file:
#     csv_reader = csv.DictReader(csv_file, delimiter=',')
#     line_count = 0
#     print(list(csv_reader))
#     # for row in csv_reader:
#     #     print(row)
#     print(f'Processed {line_count} lines.')

# ========================= write =========================

# with open('employee_file.csv', mode='a') as employee_file:
#     employee_writer = csv.writer(employee_file, delimiter=',')
#
#     employee_writer.writerow(['John Smith', 'Accounting', 'November'])
#     employee_writer.writerow(['Erica Meyers', 'IT', 'March'])
import openpyxl

path = "cars.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
max_col = sheet_obj.max_column
cars = []

for row in range(2,max_row+1):
    car_fields = ["name" , "brand" , "year" , "car_number"]
    d = {}
    for col in range(1 , max_col+1):
        cell_obj = sheet_obj.cell(row=row, column=col)
        d[car_fields[col-1]] = cell_obj.value
    cars.append(d)


with open('cars.csv', mode='w') as csv_file:
    fieldnames = ["name" , "brand" , "year" , "car_number"]
    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerows(cars)
