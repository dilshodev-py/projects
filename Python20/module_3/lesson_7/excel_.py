import openpyxl

# ===================== read ======================
# path = "cars.xlsx"
# wb_obj = openpyxl.load_workbook(path)
# sheet_obj = wb_obj.active
# print(sheet_obj.max_row)
# print(sheet_obj.max_column)
# cell_obj = sheet_obj.cell(row=2, column=1)

# print(cell_obj.value)

# ===================== write ======================


# import openpyxl module
import openpyxl

wb = openpyxl.load_workbook("cars.xlsx")


cars = [
    {
        "name": "Trailblazer",
        "brend": "chevrolet",
        "year": 2020,
        "car_number": "10A101AA"
    },
{
        "name": "Mers AMG",
        "brend": "Mers",
        "year": 2024,
        "car_number": "01A101AA"
    },
{
        "name": "BMW",
        "brend": "BMW",
        "year": 2020,
        "car_number": "20A101AA",
        "field": "20A101AA",
    }
]
sheet = wb.active
max_row = sheet.max_row
for car in cars:
    max_row += 1
    sheet.append(tuple(car.values()))

wb.save("cars.xlsx")
